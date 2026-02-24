#!/usr/bin/env python3
"""
CSV -> Excel converter (interactive input path).

- دریافت مسیر کامل (Absolute Path) فایل CSV از کاربر
- پشتیبانی از مسیرهایی که فاصله (Space) دارند
- ساخت خروجی در همان پوشه CSV
- نام خروجی: همان نام CSV با پسوند .xlsx
"""

from __future__ import annotations

import csv
import sys
from pathlib import Path

from openpyxl import Workbook


def normalize_user_path(raw: str) -> Path:
    """Normalize input path from user (handles quotes and spaces)."""
    cleaned = raw.strip().strip('"').strip("'").strip()
    return Path(cleaned)


def convert_csv_to_excel(input_csv: Path) -> tuple[Path, dict[str, int]]:
    """
    Convert input CSV to XLSX in same directory with same file stem.
    
    Returns:
        tuple: (output_xlsx_path, statistics_dict)
        statistics_dict contains:
            - total_rows: Total rows read from CSV
            - duplicates_found: Number of duplicate products skipped
            - unique_products: Number of unique products written to Excel
    """
    output_xlsx = input_csv.with_suffix(".xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    # Fixed required headers
    ws.cell(row=1, column=1, value="نام محصول")
    ws.cell(row=1, column=2, value="قیمت")

    out_row = 2
    
    # Track unique product names (case-insensitive + trimmed)
    seen_products = set()
    
    # Statistics tracking
    total_rows = 0
    duplicates_found = 0
    unique_products = 0

    with input_csv.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)

        # Skip CSV header row (if present)
        _ = next(reader, None)

        for row in reader:
            if not row:
                continue
            
            total_rows += 1

            product_name = row[0] if len(row) >= 1 else ""
            price = row[1] if len(row) >= 2 else ""

            # Skip fully empty rows only
            if product_name == "" and price == "":
                continue

            # Normalize product name for duplicate detection
            # Remove leading/trailing whitespace and convert to lowercase
            normalized_name = product_name.strip().lower()
            
            # Skip empty product names after normalization
            if not normalized_name:
                continue
            
            # Check for duplicates
            if normalized_name in seen_products:
                duplicates_found += 1
                continue
            
            # Mark this product as seen
            seen_products.add(normalized_name)

            # Preserve original values as text for exact review/edit
            ws.cell(row=out_row, column=1, value=str(product_name))
            ws.cell(row=out_row, column=2, value=str(price))
            out_row += 1
            unique_products += 1

    wb.save(output_xlsx)
    
    statistics = {
        "total_rows": total_rows,
        "duplicates_found": duplicates_found,
        "unique_products": unique_products,
    }
    
    return output_xlsx, statistics


def main() -> int:
    print("Please enter the full absolute path of the CSV file:")
    print("Example: H:\\Repo\\WordpressDevelopment\\Products-Price-Exporter\\vapeclub3-products-price.csv")
    raw_path = input("> ")

    input_csv = normalize_user_path(raw_path)

    if not input_csv.is_absolute():
        print("[ERROR] Input path must be an absolute path.", file=sys.stderr)
        return 1

    if input_csv.suffix.lower() != ".csv":
        print("[ERROR] Input file must have .csv extension.", file=sys.stderr)
        return 1

    if not input_csv.exists() or not input_csv.is_file():
        print(f"[ERROR] File not found: {input_csv}", file=sys.stderr)
        return 1

    try:
        output_xlsx, statistics = convert_csv_to_excel(input_csv)
    except Exception as exc:  # noqa: BLE001
        print(f"[ERROR] Conversion failed: {exc}", file=sys.stderr)
        return 1

    print("[OK] Conversion completed successfully.")
    print(f"[OK] Output file: {output_xlsx}")
    print("\n[STATISTICS]")
    print(f"  - Total rows processed: {statistics['total_rows']}")
    print(f"  - Unique products written: {statistics['unique_products']}")
    print(f"  - Duplicate products skipped: {statistics['duplicates_found']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
