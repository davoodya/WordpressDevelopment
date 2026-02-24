#!/usr/bin/env python3
"""
CSV to Excel Converter - Professional Edition

Features:
- OOP Architecture with extensible design
- High-performance processing for large files
- Duplicate detection (Persian/English/Mixed product names)
- Dynamic column detection (2 or 3+ columns)
- Comprehensive logging system
- Excel 2010-2026 compatibility
- UTF-8 and Persian text support

Author: Professional Python Developer
Version: 2.0.0
"""

from __future__ import annotations

import csv
import logging
import sys
import unicodedata
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Final, Iterator

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# ============================================================================
# CONFIGURATION & CONSTANTS
# ============================================================================

class Config:
    """Application configuration constants."""
    
    # File settings
    CSV_ENCODING: Final[str] = "utf-8-sig"
    EXCEL_FORMAT: Final[str] = "xlsx"
    
    # Excel compatibility (Excel 2010-2026)
    MAX_EXCEL_ROWS: Final[int] = 1_048_576
    MAX_EXCEL_COLUMNS: Final[int] = 16_384
    
    # Performance settings
    CHUNK_SIZE: Final[int] = 1000  # Process rows in chunks for memory efficiency
    
    # Logging
    LOG_FORMAT: Final[str] = "%(asctime)s - %(name)s - %(levelname)s - %(message)s"
    LOG_DATE_FORMAT: Final[str] = "%Y-%m-%d %H:%M:%S"
    
    # Headers mapping (Persian)
    DEFAULT_HEADERS: Final[dict[int, str]] = {
        0: "نام محصول",
        1: "قیمت",
        2: "دسته‌بندی"
    }


# ============================================================================
# DATA MODELS
# ============================================================================

@dataclass
class ProductRecord:
    """Represents a single product record."""
    
    product_name: str
    price: str
    category: str = ""
    row_number: int = 0
    
    def is_valid(self) -> bool:
        """Check if record has minimum required data."""
        return bool(self.product_name.strip())
    
    def normalize_name(self) -> str:
        """
        Normalize product name for duplicate detection.
        Handles Persian, English, and mixed text.
        """
        if not self.product_name:
            return ""
        
        # Strip whitespace
        normalized = self.product_name.strip()
        
        # Normalize Unicode (handle Persian variants)
        normalized = unicodedata.normalize('NFKC', normalized)
        
        # Convert to lowercase for case-insensitive comparison
        normalized = normalized.lower()
        
        # Remove extra whitespace
        normalized = ' '.join(normalized.split())
        
        return normalized


@dataclass
class ConversionStatistics:
    """Statistics for conversion process."""
    
    total_rows_read: int = 0
    empty_rows_skipped: int = 0
    invalid_rows_skipped: int = 0
    duplicate_rows_skipped: int = 0
    unique_products_written: int = 0
    columns_detected: int = 0
    processing_time_seconds: float = 0.0
    
    def __str__(self) -> str:
        """Format statistics for display."""
        return (
            f"\n{'='*60}\n"
            f"CONVERSION STATISTICS\n"
            f"{'='*60}\n"
            f"Total rows read from CSV:      {self.total_rows_read:,}\n"
            f"Empty rows skipped:            {self.empty_rows_skipped:,}\n"
            f"Invalid rows skipped:          {self.invalid_rows_skipped:,}\n"
            f"Duplicate products skipped:    {self.duplicate_rows_skipped:,}\n"
            f"Unique products written:       {self.unique_products_written:,}\n"
            f"Columns detected:              {self.columns_detected}\n"
            f"Processing time:               {self.processing_time_seconds:.2f}s\n"
            f"{'='*60}\n"
        )


# ============================================================================
# CORE BUSINESS LOGIC CLASSES
# ============================================================================

class DuplicateDetector:
    """
    High-performance duplicate detection for product names.
    Supports Persian, English, and mixed text.
    """
    
    def __init__(self) -> None:
        """Initialize duplicate detector with hash set."""
        self._seen_products: set[str] = set()
        self._duplicate_count: int = 0
    
    def is_duplicate(self, product: ProductRecord) -> bool:
        """
        Check if product name is duplicate (case-insensitive, normalized).
        
        Args:
            product: ProductRecord to check
            
        Returns:
            True if duplicate, False if unique
        """
        normalized_name = product.normalize_name()
        
        if not normalized_name:
            return False
        
        if normalized_name in self._seen_products:
            self._duplicate_count += 1
            return True
        
        self._seen_products.add(normalized_name)
        return False
    
    def get_duplicate_count(self) -> int:
        """Get total number of duplicates detected."""
        return self._duplicate_count
    
    def get_unique_count(self) -> int:
        """Get total number of unique products."""
        return len(self._seen_products)
    
    def reset(self) -> None:
        """Clear all tracked products."""
        self._seen_products.clear()
        self._duplicate_count = 0


class CSVReader:
    """
    Efficient CSV reader with chunking support for large files.
    """
    
    def __init__(self, file_path: Path, encoding: str = Config.CSV_ENCODING) -> None:
        """
        Initialize CSV reader.
        
        Args:
            file_path: Path to CSV file
            encoding: File encoding (default: UTF-8 with BOM)
        """
        self.file_path = file_path
        self.encoding = encoding
        self.logger = logging.getLogger(self.__class__.__name__)
    
    def read_records(self, skip_header: bool = True) -> Iterator[ProductRecord]:
        """
        Read CSV records efficiently with streaming.
        
        Args:
            skip_header: Whether to skip first row
            
        Yields:
            ProductRecord objects
        """
        self.logger.info(f"Reading CSV file: {self.file_path}")
        
        try:
            with self.file_path.open("r", encoding=self.encoding, newline="") as f:
                reader = csv.reader(f)
                
                # Skip header if requested
                if skip_header:
                    next(reader, None)
                
                row_number = 2 if skip_header else 1
                
                for row in reader:
                    if not row:
                        yield ProductRecord(
                            product_name="",
                            price="",
                            category="",
                            row_number=row_number
                        )
                        row_number += 1
                        continue
                    
                    # Extract fields based on available columns
                    product_name = row[0].strip() if len(row) >= 1 else ""
                    price = row[1].strip() if len(row) >= 2 else ""
                    category = row[2].strip() if len(row) >= 3 else ""
                    
                    yield ProductRecord(
                        product_name=product_name,
                        price=price,
                        category=category,
                        row_number=row_number
                    )
                    
                    row_number += 1
                    
        except Exception as e:
            self.logger.error(f"Error reading CSV file: {e}")
            raise
    
    def detect_column_count(self) -> int:
        """
        Detect number of columns in CSV file.
        
        Returns:
            Number of columns
        """
        try:
            with self.file_path.open("r", encoding=self.encoding, newline="") as f:
                reader = csv.reader(f)
                first_row = next(reader, None)
                
                if not first_row:
                    return 2  # Default to 2 columns
                
                # Count non-empty columns
                col_count = len([col for col in first_row if col.strip()])
                
                # Ensure minimum 2 columns
                return max(col_count, 2)
                
        except Exception as e:
            self.logger.warning(f"Could not detect column count: {e}. Using default (2)")
            return 2


class ExcelWriter:
    """
    High-performance Excel writer with formatting and compatibility.
    Excel 2010-2026 compatible.
    """
    
    def __init__(self, output_path: Path, column_count: int = 2) -> None:
        """
        Initialize Excel writer.
        
        Args:
            output_path: Path for output Excel file
            column_count: Number of columns to create
        """
        self.output_path = output_path
        self.column_count = column_count
        self.logger = logging.getLogger(self.__class__.__name__)
        
        # Create workbook
        self.workbook = Workbook()
        self.worksheet: Worksheet = self.workbook.active
        self.worksheet.title = "Products"
        
        # Initialize
        self._setup_headers()
        self._current_row = 2
    
    def _setup_headers(self) -> None:
        """Setup Excel headers with formatting."""
        self.logger.debug(f"Setting up {self.column_count} column headers")
        
        # Header font style
        header_font = Font(bold=True, size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers based on column count
        for col_idx in range(self.column_count):
            header_text = Config.DEFAULT_HEADERS.get(col_idx, f"ستون {col_idx + 1}")
            cell = self.worksheet.cell(row=1, column=col_idx + 1, value=header_text)
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Auto-adjust column widths
        for col_idx in range(self.column_count):
            column_letter = get_column_letter(col_idx + 1)
            self.worksheet.column_dimensions[column_letter].width = 30
    
    def write_record(self, record: ProductRecord) -> None:
        """
        Write a product record to Excel.
        
        Args:
            record: ProductRecord to write
        """
        if self._current_row > Config.MAX_EXCEL_ROWS:
            self.logger.warning(f"Reached Excel row limit: {Config.MAX_EXCEL_ROWS}")
            raise ValueError(f"Excel row limit reached: {Config.MAX_EXCEL_ROWS}")
        
        # Write product name
        self.worksheet.cell(
            row=self._current_row,
            column=1,
            value=record.product_name
        )
        
        # Write price
        self.worksheet.cell(
            row=self._current_row,
            column=2,
            value=record.price
        )
        
        # Write category if present
        if self.column_count >= 3 and record.category:
            self.worksheet.cell(
                row=self._current_row,
                column=3,
                value=record.category
            )
        
        self._current_row += 1
    
    def save(self) -> None:
        """Save Excel file to disk."""
        self.logger.info(f"Saving Excel file: {self.output_path}")
        
        try:
            self.workbook.save(self.output_path)
            self.logger.info(f"Excel file saved successfully: {self.output_path}")
        except Exception as e:
            self.logger.error(f"Error saving Excel file: {e}")
            raise
    
    def get_rows_written(self) -> int:
        """Get number of data rows written (excluding header)."""
        return self._current_row - 2


class PathValidator:
    """Validator for file paths and user input."""
    
    @staticmethod
    def normalize_user_path(raw: str) -> Path:
        """
        Normalize input path from user (handles quotes and spaces).
        
        Args:
            raw: Raw input string from user
            
        Returns:
            Normalized Path object
        """
        cleaned = raw.strip().strip('"').strip("'").strip()
        return Path(cleaned)
    
    @staticmethod
    def validate_csv_path(path: Path) -> tuple[bool, str]:
        """
        Validate CSV file path.
        
        Args:
            path: Path to validate
            
        Returns:
            Tuple of (is_valid, error_message)
        """
        if not path.is_absolute():
            return False, "Input path must be an absolute path"
        
        if path.suffix.lower() != ".csv":
            return False, "Input file must have .csv extension"
        
        if not path.exists():
            return False, f"File not found: {path}"
        
        if not path.is_file():
            return False, f"Path is not a file: {path}"
        
        return True, ""


# ============================================================================
# MAIN CONVERTER CLASS
# ============================================================================

class CSVToExcelConverter:
    """
    Main converter class - orchestrates the conversion process.
    Designed for extensibility (batch processing, record selection, etc.)
    """
    
    def __init__(self, input_path: Path, output_path: Path | None = None) -> None:
        """
        Initialize converter.
        
        Args:
            input_path: Path to input CSV file
            output_path: Optional custom output path (default: same as input with .xlsx)
        """
        self.input_path = input_path
        self.output_path = output_path or input_path.with_suffix(".xlsx")
        
        # Setup logging
        self.logger = logging.getLogger(self.__class__.__name__)
        
        # Components
        self.csv_reader = CSVReader(self.input_path)
        self.duplicate_detector = DuplicateDetector()
        self.statistics = ConversionStatistics()
        
        # Performance tracking
        self._start_time: float = 0.0
    
    def convert(self) -> ConversionStatistics:
        """
        Execute the conversion process.
        
        Returns:
            ConversionStatistics object with results
        """
        import time
        self._start_time = time.time()
        
        self.logger.info("="*60)
        self.logger.info("Starting CSV to Excel conversion")
        self.logger.info(f"Input:  {self.input_path}")
        self.logger.info(f"Output: {self.output_path}")
        self.logger.info("="*60)
        
        try:
            # Detect column count
            self.statistics.columns_detected = self.csv_reader.detect_column_count()
            self.logger.info(f"Detected {self.statistics.columns_detected} columns in CSV")
            
            # Initialize Excel writer
            excel_writer = ExcelWriter(self.output_path, self.statistics.columns_detected)
            
            # Process records
            self._process_records(excel_writer)
            
            # Save Excel file
            excel_writer.save()
            
            # Finalize statistics
            self.statistics.processing_time_seconds = time.time() - self._start_time
            self.statistics.unique_products_written = excel_writer.get_rows_written()
            self.statistics.duplicate_rows_skipped = self.duplicate_detector.get_duplicate_count()
            
            self.logger.info("Conversion completed successfully")
            self.logger.info(str(self.statistics))
            
            return self.statistics
            
        except Exception as e:
            self.logger.error(f"Conversion failed: {e}", exc_info=True)
            raise
    
    def _process_records(self, excel_writer: ExcelWriter) -> None:
        """
        Process CSV records and write to Excel.
        
        Args:
            excel_writer: ExcelWriter instance
        """
        for record in self.csv_reader.read_records(skip_header=True):
            self.statistics.total_rows_read += 1
            
            # Skip empty rows
            if not record.product_name and not record.price:
                self.statistics.empty_rows_skipped += 1
                self.logger.debug(f"Skipped empty row: {record.row_number}")
                continue
            
            # Validate record
            if not record.is_valid():
                self.statistics.invalid_rows_skipped += 1
                self.logger.warning(f"Skipped invalid row {record.row_number}: empty product name")
                continue
            
            # Check for duplicates
            if self.duplicate_detector.is_duplicate(record):
                self.logger.debug(
                    f"Skipped duplicate product at row {record.row_number}: "
                    f"{record.product_name}"
                )
                continue
            
            # Write to Excel
            excel_writer.write_record(record)
            
            # Log progress for large files (every 1000 records)
            if self.statistics.total_rows_read % 1000 == 0:
                self.logger.info(f"Processed {self.statistics.total_rows_read:,} rows...")


# ============================================================================
# LOGGING SETUP
# ============================================================================

class LoggerSetup:
    """Setup comprehensive logging system."""
    
    @staticmethod
    def setup_logging(log_file_path: Path, verbose: bool = False) -> None:
        """
        Configure logging with file and console handlers.
        
        Args:
            log_file_path: Path for log file
            verbose: Enable verbose console output
        """
        # Root logger
        root_logger = logging.getLogger()
        root_logger.setLevel(logging.DEBUG)
        
        # Clear existing handlers
        root_logger.handlers.clear()
        
        # File handler (detailed logs)
        file_handler = logging.FileHandler(
            log_file_path,
            mode="w",
            encoding="utf-8"
        )
        file_handler.setLevel(logging.DEBUG)
        file_formatter = logging.Formatter(
            Config.LOG_FORMAT,
            datefmt=Config.LOG_DATE_FORMAT
        )
        file_handler.setFormatter(file_formatter)
        root_logger.addHandler(file_handler)
        
        # Console handler (user-friendly output)
        console_handler = logging.StreamHandler(sys.stdout)
        console_handler.setLevel(logging.DEBUG if verbose else logging.INFO)
        console_formatter = logging.Formatter("%(message)s")
        console_handler.setFormatter(console_formatter)
        root_logger.addHandler(console_handler)


# ============================================================================
# USER INTERFACE
# ============================================================================

class ConsoleUI:
    """Console user interface for interactive mode."""
    
    @staticmethod
    def print_banner() -> None:
        """Print application banner."""
        print("\n" + "="*60)
        print("CSV to Excel Converter - Professional Edition v2.0.0")
        print("="*60 + "\n")
    
    @staticmethod
    def get_csv_path() -> Path:
        """
        Prompt user for CSV file path.
        
        Returns:
            Validated Path object
        """
        print("Please enter the full absolute path of the CSV file:")
        print("Example: H:\\Repo\\WordpressDevelopment\\Products-Price-Exporter\\vapeclub3-products-price.csv")
        print()
        
        raw_path = input("> ").strip()
        return PathValidator.normalize_user_path(raw_path)
    
    @staticmethod
    def print_error(message: str) -> None:
        """Print error message."""
        print(f"\n[ERROR] {message}\n", file=sys.stderr)
    
    @staticmethod
    def print_success(output_path: Path, statistics: ConversionStatistics) -> None:
        """Print success message with statistics."""
        print("\n" + "="*60)
        print("[SUCCESS] Conversion completed successfully!")
        print("="*60)
        print(f"\nOutput file: {output_path}")
        print(statistics)


# ============================================================================
# MAIN ENTRY POINT
# ============================================================================

def main() -> int:
    """
    Main entry point for the application.
    
    Returns:
        Exit code (0 for success, 1 for error)
    """
    # Print banner
    ConsoleUI.print_banner()
    
    # Get input path
    try:
        input_csv = ConsoleUI.get_csv_path()
    except KeyboardInterrupt:
        print("\n\n[INFO] Operation cancelled by user.")
        return 0
    
    # Validate path
    is_valid, error_message = PathValidator.validate_csv_path(input_csv)
    if not is_valid:
        ConsoleUI.print_error(error_message)
        return 1
    
    # Setup logging
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = input_csv.parent / f"conversion_{timestamp}.log"
    LoggerSetup.setup_logging(log_file, verbose=False)
    
    logger = logging.getLogger("Main")
    logger.info(f"Log file created: {log_file}")
    
    # Execute conversion
    try:
        converter = CSVToExcelConverter(input_csv)
        statistics = converter.convert()
        
        # Print results
        ConsoleUI.print_success(converter.output_path, statistics)
        
        print(f"Log file: {log_file}\n")
        
        return 0
        
    except KeyboardInterrupt:
        logger.info("Operation cancelled by user")
        print("\n\n[INFO] Operation cancelled by user.")
        return 0
        
    except Exception as exc:
        logger.error(f"Conversion failed: {exc}", exc_info=True)
        ConsoleUI.print_error(f"Conversion failed: {exc}")
        print(f"Check log file for details: {log_file}\n")
        return 1


if __name__ == "__main__":
    sys.exit(main())
