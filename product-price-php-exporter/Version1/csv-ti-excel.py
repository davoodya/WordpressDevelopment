#!/usr/bin/env python3
"""
Convert a CSV file (product name + price) to an editable Excel (.xlsx) file.

Requirements covered:
- Reads CSV with UTF-8 BOM support (utf-8-sig)
- Creates Excel with exactly 2 columns:
  1) نام محصول
  2) قیمت
- Copies all rows as-is from CSV into Excel
- Keeps value text exactly as read from CSV (including empty price cells)
- Generates .xlsx compatible with Microsoft Office 2010+ (2010..2026)
"""

from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

from openpyxl import Workbook



print("Please enter the full absolute path of the CSV file:")
print("Example: H:\\Repo\\WordpressDevelopment\\Products-Price-Exporter\\vapeclub3-products-price.csv")
print()

DEFAULT_INPUT = input("> ").strip()
#DEFAULT_INPUT = Path(r"H:\Repo\WordpressDevelopment\Products-Price-Exporter\vapeclub3-products-price.csv")

def normalize_user_path(raw: str) -> Path:
        """
        Normalize input path from user (handles quotes and spaces).
        
        Args:
            raw: Raw input string from user
            
        Returns:
            Normalized Path object
        """
        cleaned = raw.strip().strip('"').strip("'").strip()
        return Path(cleaned)
        
DEFAULT_INPUT = normalize_user_path(DEFAULT_INPUT)


def validate_csv_path(path: Path) -> tuple[bool, str]:
        """
        Validate CSV file path.
        
        Args:
            path: Path to validate
            
        Returns:
            Tuple of (is_valid, error_message)
        """
        if not path.is_absolute():
            return False, "Input path must be an absolute path"
        
        if path.suffix.lower() != ".csv":
            return False, "Input file must have .csv extension"
        
        if not path.exists():
            return False, f"File not found: {path}"
        
        if not path.is_file():
            return False, f"Path is not a file: {path}"
        
        return True, ""


is_valid, error_message = validate_csv_path(DEFAULT_INPUT)
if not is_valid:
    print(error_message)
    exit(1)
    
DEFAULT_OUTPUT = DEFAULT_INPUT.with_suffix(".xlsx")

def convert_csv_to_excel(input_csv: Path, output_xlsx: Path) -> tuple[int, int]:
    """
    Convert CSV -> XLSX.

    Returns:
        (rows_written, skipped_empty_rows)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    # Required fixed output headers
    ws.cell(row=1, column=1, value="نام محصول")
    ws.cell(row=1, column=2, value="قیمت")

    rows_written = 0
    skipped_empty_rows = 0
    out_row = 2

    with input_csv.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)

        # Skip CSV header if present. We still force our own fixed header in Excel.
        first = next(reader, None)
        if first is None:
            wb.save(output_xlsx)
            return rows_written, skipped_empty_rows

        for row in reader:
            # Normalize row length to exactly 2 columns
            if len(row) == 0:
                skipped_empty_rows += 1
                continue

            product_name = row[0] if len(row) >= 1 else ""
            price = row[1] if len(row) >= 2 else ""

            # Skip totally blank rows only
            if product_name == "" and price == "":
                skipped_empty_rows += 1
                continue

            # Write as text to preserve exact CSV values
            ws.cell(row=out_row, column=1, value=str(product_name))
            ws.cell(row=out_row, column=2, value=str(price))

            out_row += 1
            rows_written += 1

    wb.save(output_xlsx)
    return rows_written, skipped_empty_rows


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Convert products-price CSV to editable Excel file (.xlsx)."
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_INPUT,
        help=f"Input CSV path (default: {DEFAULT_INPUT})",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"Output XLSX path (default: {DEFAULT_OUTPUT})",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    input_csv = args.input
    output_xlsx = args.output

    if not input_csv.exists():
        print(f"[ERROR] Input CSV not found: {input_csv}", file=sys.stderr)
        return 1

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)

    rows_written, skipped_empty_rows = convert_csv_to_excel(input_csv, output_xlsx)

    print(f"[OK] Excel file created: {output_xlsx}")
    print(f"[INFO] Rows written: {rows_written}")
    print(f"[INFO] Empty rows skipped: {skipped_empty_rows}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
